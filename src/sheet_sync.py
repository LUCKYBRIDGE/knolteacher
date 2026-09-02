import os
import sys
import csv
import json
import urllib.request
import urllib.parse
from typing import Optional, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class SheetSyncManager:
    """
    엑셀(.xlsx) 및 구글 시트(Google Sheets) 시간표/일과표 연동 관리자
    """
    GOOGLE_SHEET_GUIDE_TEXT = """
[구글 시트 연동 방법]
1. 구글 스프레드시트에서 [파일] -> [공유] -> [웹에 게시]를 클릭합니다.
2. '전체 문서'를 '쉼표로 구분된 값(.csv)' 형식으로 선택하고 [게시]를 누릅니다.
3. 생성된 URL 링크를 아래 입력창에 붙여넣고 [동기화]를 누르면 실시간으로 시간표가 연동됩니다!
    """.strip()

    @staticmethod
    def export_to_excel(filepath: str, weekly_data: dict[str, list[dict[str, str]]], periods_data: list[dict[str, Any]]) -> bool:
        """시간표 및 일과표를 엑셀 파일로 내보내기"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "학급시간표"

            # 스타일 정의
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            lunch_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            lunch_font = Font(name="맑은 고딕", size=10, bold=True, color="92400E")
            default_font = Font(name="맑은 고딕", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1")
            )

            # 제목
            ws.merge_cells("A1:G1")
            ws["A1"] = "🏫 2026학년도 학급 시간표 및 일과표"
            ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True, color="1E293B")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35

            # 헤더
            headers = ["교시", "수업 시간", "월요일", "화요일", "수요일", "목요일", "금요일"]
            for col_num, h_text in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=h_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[3].height = 26

            # 내용 채우기
            day_keys = ["mon", "tue", "wed", "thu", "fri"]
            row_idx = 4
            p_lesson_idx = 0

            for p in periods_data:
                name = p.get("name", "")
                time_range = f"{p.get('start', '')} ~ {p.get('end', '')}"
                is_lunch = p.get("is_lunch", False)

                ws.cell(row=row_idx, column=1, value=name).font = Font(name="맑은 고딕", size=10, bold=True)
                ws.cell(row=row_idx, column=2, value=time_range).font = default_font

                if is_lunch:
                    ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=7)
                    lunch_cell = ws.cell(row=row_idx, column=3, value="🍱 점심식사 및 휴식")
                    lunch_cell.fill = lunch_fill
                    lunch_cell.font = lunch_font
                    lunch_cell.alignment = center_align
                else:
                    for col_idx, d_key in enumerate(day_keys, 3):
                        subjects_list = weekly_data.get(d_key, [])
                        subj_obj = subjects_list[p_lesson_idx] if p_lesson_idx < len(subjects_list) else {"subject": "", "tag": "담임"}
                        
                        subj_name = subj_obj.get("subject", "") if isinstance(subj_obj, dict) else str(subj_obj)
                        tag = subj_obj.get("tag", "담임") if isinstance(subj_obj, dict) else "담임"

                        disp_text = f"{subj_name} [{tag}]" if tag != "담임" and subj_name else subj_name
                        cell = ws.cell(row=row_idx, column=col_idx, value=disp_text)
                        cell.font = default_font
                        cell.alignment = center_align

                    p_lesson_idx += 1

                for c in range(1, 8):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.border = thin_border
                    if not is_lunch:
                        cell.alignment = center_align

                ws.row_dimensions[row_idx].height = 24
                row_idx += 1

            # 열 너비 자동 조정
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 18
            for c_letter in ["C", "D", "E", "F", "G"]:
                ws.column_dimensions[c_letter].width = 16

            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Excel export error: {e}")
            return False

    @staticmethod
    def import_from_excel_or_csv(filepath: str) -> tuple[bool, Optional[dict[str, list[dict[str, str]]]], str]:
        """엑셀 또는 CSV 파일에서 주간 시간표 불러오기"""
        if not os.path.exists(filepath):
            return False, None, "파일이 존재하지 않습니다."

        day_keys = ["mon", "tue", "wed", "thu", "fri"]
        parsed_data = {k: [] for k in day_keys}

        try:
            if filepath.endswith(".xlsx"):
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
                # 4행부터 데이터 시작 (헤더는 3행)
                for row in ws.iter_rows(min_row=4, max_col=7, values_only=True):
                    if not row or not row[0]:
                        continue
                    name_str = str(row[0])
                    if "점심" in name_str:
                        continue
                    for c_idx, d_key in enumerate(day_keys, 2):
                        val = str(row[c_idx] or "").strip()
                        # 태그 분리 (예: "국어 [전담]")
                        tag = "담임"
                        if "[전담]" in val:
                            tag = "전담"
                            val = val.replace("[전담]", "").strip()
                        elif "[외강]" in val or "[외부강사]" in val:
                            tag = "외강"
                            val = val.replace("[외강]", "").replace("[외부강사]", "").strip()

                        parsed_data[d_key].append({"subject": val, "tag": tag})
                return True, parsed_data, "엑셀 시간표를 성공적으로 가져왔습니다."

            elif filepath.endswith(".csv"):
                with open(filepath, "r", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    for r in rows[1:]:
                        if not r or "점심" in r[0]:
                            continue
                        for c_idx, d_key in enumerate(day_keys, 2):
                            if c_idx < len(r):
                                val = r[c_idx].strip()
                                tag = "담임"
                                if "[전담]" in val:
                                    tag = "전담"
                                    val = val.replace("[전담]", "").strip()
                                elif "[외강]" in val or "[외부강사]" in val:
                                    tag = "외강"
                                    val = val.replace("[외강]", "").replace("[외부강사]", "").strip()
                                parsed_data[d_key].append({"subject": val, "tag": tag})
                return True, parsed_data, "CSV 시간표를 성공적으로 가져왔습니다."
            else:
                return False, None, "지원하지 않는 파일 형식입니다 (.xlsx, .csv 가능)."
        except Exception as e:
            return False, None, f"파일 읽기 오류: {str(e)}"

    @staticmethod
    def sync_from_google_sheet_csv(url: str) -> tuple[bool, Optional[dict[str, list[dict[str, str]]]], str]:
        """구글 스프레드시트 공개 웹 CSV URL로부터 시간표 동기화"""
        if not url.strip().startswith("http"):
            return False, None, "올바른 구글 시트 웹 게시 URL을 입력해주세요."

        day_keys = ["mon", "tue", "wed", "thu", "fri"]
        parsed_data = {k: [] for k in day_keys}

        try:
            req = urllib.request.Request(url.strip(), headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=6) as response:
                content = response.read().decode("utf-8-sig")

            reader = csv.reader(content.splitlines())
            rows = list(reader)

            if len(rows) < 2:
                return False, None, "구글 시트 데이터가 비어있습니다."

            for r in rows:
                if not r:
                    continue
                # 교시 행 판별
                first_cell = r[0].strip()
                if "점심" in first_cell or "교시" not in first_cell and not first_cell.isdigit():
                    continue

                for c_idx, d_key in enumerate(day_keys, 1):
                    if c_idx < len(r):
                        val = r[c_idx].strip()
                        tag = "담임"
                        if "[전담]" in val:
                            tag = "전담"
                            val = val.replace("[전담]", "").strip()
                        elif "[외강]" in val:
                            tag = "외강"
                            val = val.replace("[외강]", "").strip()
                        parsed_data[d_key].append({"subject": val, "tag": tag})

            return True, parsed_data, "구글 시트로부터 시간표를 성공적으로 동기화했습니다!"
        except Exception as e:
            return False, None, f"구글 시트 연동 실패: {str(e)}"

sheet_sync_manager = SheetSyncManager()
