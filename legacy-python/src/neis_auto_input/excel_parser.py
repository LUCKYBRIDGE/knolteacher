import os
import openpyxl
from typing import Optional, Any

class ExcelNeisParser:
    """
    나이스 자동입력을 위한 엑셀 파서
    - 시트 목록 확인
    - 번호, 이름, 내용 열 자동 감지 및 사용자 지정 매핑
    """
    def __init__(self):
        self.workbook: Optional[openpyxl.Workbook] = None
        self.filepath: str = ""
        self.sheet_names: list[str] = []
        self.active_sheet_name: str = ""
        self.headers: list[str] = []
        self.columns_count: int = 0

    def load_file(self, filepath: str) -> tuple[bool, str, list[str]]:
        if not os.path.exists(filepath):
            return False, "파일이 존재하지 않습니다.", []
        if not filepath.endswith(".xlsx"):
            return False, "xlsx 형식의 엑셀 파일만 지원합니다.", []

        try:
            self.filepath = filepath
            self.workbook = openpyxl.load_workbook(filepath, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            if not self.sheet_names:
                return False, "엑셀 파일에 시트가 없습니다.", []
            self.active_sheet_name = self.sheet_names[0]
            return True, "엑셀 파일을 성공적으로 불러왔습니다.", self.sheet_names
        except Exception as e:
            return False, f"엑셀 파일 열기 실패: {str(e)}", []

    def select_sheet(self, sheet_name: str) -> tuple[bool, str, list[str], dict[str, int]]:
        """
        시트 선택 후 헤더 목록과 자동 감지된 열 인덱스(0-based) 반환
        반환: (성공여부, 메시지, 헤더리스트, 감지된열맵{'num_col': int, 'name_col': int, 'content_col': int})
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return False, "유효하지 않은 시트입니다.", [], {}

        self.active_sheet_name = sheet_name
        ws = self.workbook[sheet_name]

        # 1~5행 중에서 가장 열이 많은 행을 헤더 행으로 탐색
        header_row_idx = 1
        best_headers = []

        for r_idx in range(1, min(6, ws.max_row + 1)):
            row_vals = [str(c.value or "").strip() for c in ws[r_idx]]
            non_empty_count = sum(1 for v in row_vals if v)
            if non_empty_count > len(best_headers):
                best_headers = row_vals
                header_row_idx = r_idx

        # 빈 끝 열 정리
        while best_headers and not best_headers[-1]:
            best_headers.pop()

        self.headers = [h if h else f"열 {i+1}" for i, h in enumerate(best_headers)]
        self.columns_count = len(self.headers)

        # 열 자동 감지
        detected = self._auto_detect_columns(self.headers, ws, header_row_idx)
        return True, f"시트 '{sheet_name}' (헤더: {len(self.headers)}개 열) 로드 완료", self.headers, detected

    def _auto_detect_columns(self, headers: list[str], ws, header_row_idx: int) -> dict[str, int]:
        detected = {"num_col": -1, "name_col": -1, "content_col": -1}

        # 1. 헤더 이름 기반 매칭
        for idx, h in enumerate(headers):
            h_clean = h.replace(" ", "").lower()
            if detected["num_col"] == -1 and any(k in h_clean for k in ["번호", "학생번호", "no", "num", "번", "순번"]):
                detected["num_col"] = idx
            elif detected["name_col"] == -1 and any(k in h_clean for k in ["성명", "이름", "학생명", "학생성명", "name"]):
                detected["name_col"] = idx
            elif detected["content_col"] == -1 and any(k in h_clean for k in ["행동특성", "종합의견", "행발", "의견", "내용", "평어", "평가", "특기사항"]):
                detected["content_col"] = idx

        # 2. 감지 안 된 경우 데이터 샘플 분석
        sample_rows = list(ws.iter_rows(min_row=header_row_idx+1, max_row=min(header_row_idx+10, ws.max_row), values_only=True))
        if sample_rows:
            col_types = []
            for col_i in range(len(headers)):
                vals = [r[col_i] for r in sample_rows if col_i < len(r) and r[col_i] is not None]
                is_mostly_int = False
                avg_len = 0
                if vals:
                    int_count = sum(1 for v in vals if str(v).strip().isdigit())
                    is_mostly_int = (int_count / len(vals)) >= 0.7
                    avg_len = sum(len(str(v)) for v in vals) / len(vals)
                col_types.append((is_mostly_int, avg_len))

            if detected["num_col"] == -1:
                for col_i, (is_int, _) in enumerate(col_types):
                    if is_int:
                        detected["num_col"] = col_i
                        break

            if detected["content_col"] == -1:
                # 가장 긴 텍스트를 가진 열을 내용 열로 추정
                longest_col = max(range(len(col_types)), key=lambda c: col_types[c][1])
                if longest_col != detected["num_col"] and longest_col != detected["name_col"]:
                    detected["content_col"] = longest_col

            if detected["name_col"] == -1:
                for col_i in range(len(headers)):
                    if col_i != detected["num_col"] and col_i != detected["content_col"]:
                        detected["name_col"] = col_i
                        break

        # 기본값 폴백 (순서대로 0, 1, 2)
        if detected["num_col"] == -1 and len(headers) > 0: detected["num_col"] = 0
        if detected["name_col"] == -1 and len(headers) > 1: detected["name_col"] = 1
        if detected["content_col"] == -1 and len(headers) > 2: detected["content_col"] = 2

        return detected

    def parse_data(self, num_col: int, name_col: int, content_col: int, header_row_idx: int = 1) -> list[dict[str, Any]]:
        """
        엑셀 행 데이터를 파싱하여 학생 리스트 반환
        """
        if not self.workbook or not self.active_sheet_name:
            return []

        ws = self.workbook[self.active_sheet_name]
        records = []

        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx+1, values_only=True), start=header_row_idx+1):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            num_val = row[num_col] if num_col < len(row) else None
            name_val = row[name_col] if name_col >= 0 and name_col < len(row) else ""
            content_val = row[content_col] if content_col < len(row) else ""

            records.append({
                "row_index": r_idx,
                "raw_number": num_val,
                "name": str(name_val or "").strip(),
                "content": str(content_val or "").strip()
            })

        return records
